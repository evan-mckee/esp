"""
ESP Scrum Manager 1.1
7/12/2021
Evan McKee
"""


import random
import string
import json
from datetime import datetime
from benedict import benedict  # pip install python-benedict
from openpyxl import Workbook  # Remove if not printing to Excel
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font, NamedStyle
from openpyxl.utils import get_column_letter

class EspApp:
    """
    Evan-Style-Python or ESP:
    Decouple the input variables and data from your code and store them in an external JSON file.
    Read the file into a dict at the beginning; update during runtime, save and load whenever.

    Scrum Manager:
    Scrum boards are used to organize to-do lists into projects, stories, and tasks.
    Data is sorted into Projects, which contain Stories, which contain Tasks.

    During use, there are three ways to manage resources: The JSON file can be edited outside of
    Python, the console interface (which is itself described in the JSON file) can update
    resources, or resources can be loaded in by a text file.

    This project is meant as a demonstration of resource management and HMI using ESP, and not as
    a standalone app.
    """

    def __init__(self, esp_data, output_file):
        """
        Read the JSON file and initialize runtime variables.

        :param esp_data: Path to JSON file to read (str)
        :param output_file: Path to JSON file to write (str)
        """

        self.d = self.read_esp_data(esp_data)

        self.active_project = ''
        self.active_story = ''
        self.active_task = ''
        self.active_attributes = ''
        self.selected_attribute = ''
        self.active_resource = ''
        self.active_filter = ''
        self.window_chain = []
        self.input_text = ''
        self.time_format = '%Y_%m_%d_%H_%M_%S'

        self.output_file = output_file

    def run(self):
        """
        Step through text-based menus to add, delete, and edit scrum board notes.

        Note: Parameters for each window are stored in external JSON file.
        """
        current_window = 'home'
        while current_window != 'exit':
            window = self.d['windows'][current_window]
            exitnum = '9'
            valid_options = [exitnum]
            choice_vars = []

            # Print menu prompt.
            if window['prompt_type'] == "static":
                print('')
                print(window['prompt'])
                print('-' * len(window['prompt']))
            elif window['prompt_type'] == 'dynamic':
                print('')
                txt = self.get_replacements(window['prompt'], window['replacements'])
                print(txt)
                print('-' * len(txt))

            # Gather input options and labels.
            if window['choice_type'] == 'static_numeric':  # static_numeric: Fixed choices, selected by numeric input.
                exitnum = self.get_exitnum(window['choices'])
                valid_options = [str(k) for k in list(range(0, len(window['choices'])))]
                valid_options.append(exitnum)

                self.print_menu([k[0] for k in window['choices']], exitnum)
            elif window['choice_type'] == 'dynamic_numeric':  # dynamic_numeric: Options sourced from window['source'].
                if 'runtime.' in window['source']:
                    choice_vars = getattr(self, window['source'].replace('runtime.', ''))
                else:
                    choice_vars = self.d[self.replace_at_symbol(window['source'])]
                if window['filter'] != '':
                    choice_vars = [k for k in choice_vars if getattr(self, window['filter']) in k]
                exitnum = self.get_exitnum(choice_vars)
                valid_options = [str(k) for k in list(range(0, len(choice_vars)))]
                valid_options.append(exitnum)
                choice_labels = []
                if window['labels'] == "":
                    choice_labels = choice_vars
                else:
                    for k in choice_vars:
                        choice_labels.append(self.d[k][window['labels']])
                self.print_menu(choice_labels, exitnum)
            elif window['choice_type'] == 'dynamic_paths':  # dynamic_paths: Each option is a dot path to a resource.
                true_paths = []
                named_paths = []
                for k in window['filters']:
                    true_paths.extend(self.d[k])
                true_paths.sort()
                resource_text = {}
                for k in true_paths:
                    if k.split('.')[-1] not in resource_text:
                        resource_text[k.split('.')[-1]] = self.d[k + '.text']
                for k in true_paths:
                    txt = str(k)
                    for n in resource_text:
                        txt = txt.replace(n, resource_text[n])
                    named_paths.append(txt)
                named_paths = [k.replace('projects.', '').replace('storys.', '').replace('tasks.', '')
                               for k in named_paths]
                choice_vars = true_paths
                exitnum = self.get_exitnum(choice_vars)
                valid_options = [str(k) for k in list(range(0, len(choice_vars)))]
                valid_options.append(exitnum)
                choice_labels = named_paths
                self.print_menu(choice_labels, exitnum)
            elif window['choice_type'] == 'input_chain':  # input_chain: Offer a series of prompts, store typed input.
                for k in window['choices']:
                    print(k[0])
                    x = input()
                    setattr(self, k[1], x)
                self.do_window_function(window['function'])
                current_window = window['next_window']

            # Read and decode input.
            if window['choice_type'] in ['static_numeric', 'dynamic_numeric', 'dynamic_paths']:
                x = ''
                while x not in valid_options:
                    x = input()
                if x == exitnum:
                    current_window = 'exit'
                else:
                    if window['choice_type'] == 'static_numeric':
                        selected = window['choices'][int(x)]
                        if selected[1] != "":
                            self.do_window_function(selected[1])
                        if selected[2] != "":
                            self.window_chain = selected[2].split(',')
                        current_window = self.window_chain[0]
                        self.window_chain.pop(0)
                    elif window['choice_type'] == 'dynamic_numeric':
                        selected = choice_vars[int(x)]
                        setattr(self, window['write_to'], selected)
                        current_window = self.window_chain[0]
                        self.window_chain.pop(0)
                    elif window['choice_type'] == 'dynamic_paths':
                        selected = choice_vars[int(x)]
                        setattr(self, window['write_to'], selected)
                        if window['function'] != "":
                            self.do_window_function(window['function'])
                        current_window = window['next_window']

        # Save JSON file and exit.
        self.write_esp_data(self.output_file)

    # Input Methods
    def read_esp_data(self, data_file):
        """
        Read JSON data into a dict.

        Note: python-benedict allows dict indexing by dot notation (my_dict['a']['b'] == my_dict['a.b'].
        While not required, is useful for managing nested dicts called dynamically.

        :param data_file: Path to JSON file (str)
        """
        with open(data_file, 'r') as f:
            return benedict(json.load(f))

    def read_resource_loader(self, txt_file):
        """
        Read resources in from a text file (see example_resource_loader.txt)

        :param txt_file: Path to resource loader file (str)
        """
        with open(txt_file, 'r') as f:
            lines = [k.replace('\n', '') for k in f.readlines() if k != '\n' and k.startswith('#') is False]
            project = ''
            story = ''
            task = ''
            status = ''
            statuses = {
                'T': 'ToDo',
                'I': 'In Progress',
                'R': 'Review',
                'B': 'Blocked',
                'C': 'Complete',
            }
            for line in lines:
                # Tasks
                if line.startswith('\t\t'):
                    line = line.replace('\t\t','')
                    if line[-4:-1] == ' - ':
                        status = statuses[line[-1]]
                        line = line[:-4]
                    else:
                        status = 'TODO'
                    if line.startswith(':'):
                        task = self.get_task(line.replace(':', ''))
                        self.d[task]['status'] = status
                    else:
                        self.add_task(project, story, line, status)
                # Stories
                elif line.startswith('\t'):
                    line = line.replace('\t','')
                    if line.startswith(':'):
                        story = self.get_story(project, line.replace(':', ''))
                    else:
                        self.add_story(project, line)
                        story = self.get_story(project, line)
                # Projects
                else:
                    if line.startswith(':'):
                        project = self.get_project(line.replace(':', ''))
                    else:
                        self.add_project(line)
                        project = self.get_project(line)

    # Process Methods
    def do_window_function(self, function):
        """
        Perform a dynamically chosen function specified by a dict value.

        :param function: Key to available function in this method (str)
        """
        if function == 'print_tree':
            for k in self.d['projects']:
                print(self.d['.'.join(['projects', k, 'text'])])
                for j in self.d['.'.join(['projects', k, 'storys'])]:
                    print('\t' + self.d['.'.join(['projects', k, 'storys', j, 'text'])])
                    for n in self.d['.'.join(['projects', k, 'storys', j, 'tasks'])]:
                        print('\t\t' + self.d['.'.join(['projects', k, 'storys', j, 'tasks', n, 'status'])].upper() +
                              ': ' + self.d['.'.join(['projects', k, 'storys', j, 'tasks', n, 'text'])] + 
                              ' ' + self.d['.'.join(['projects', k, 'storys', j, 'tasks', n, 'notes'])])
        elif function == 'scrum_projects':
            self.print_scrum_board(self.d['project_paths'])
        elif function == 'add_project':
            self.add_project(self.input_text)
        elif function == 'add_story':
            self.add_story(self.active_project, self.input_text)
        elif function == 'add_task':
            self.add_task(self.active_project, self.active_story, self.input_text)
        elif function == 'del_project':
            self.del_project(self.active_project)
        elif function == 'del_story':
            self.del_story(self.active_story)
        elif function == 'del_task':
            self.del_task(self.active_task)
        elif function == 'set_active_attributes':
            if self.active_resource.count('.') == 1:
                self.active_attributes = ['text']
            elif self.active_resource.count('.') == 3:
                self.active_attributes = ['text']
            elif self.active_resource.count('.') == 5:
                self.active_attributes = ['text', 'status', 'notes']
            self.window_chain = ['update_attribute', 'home']
        elif function == 'update_attribute':
            self.d[self.active_resource + '.' + self.selected_attribute] = self.input_text
            time = datetime.now().strftime(self.time_format)
            self.d['log'].append('Updated ' + self.active_resource + '.' + self.selected_attribute + ' to '
                                 + self.input_text + ' at ' + time)
        elif function == 'print_to_excel':
            self.print_to_excel(self.d['project_paths'], 'SCRUM.xlsx')
        elif function == 'read_from_text':
            print('Text file? (Include extension)')
            x = input()
            self.read_resource_loader(x)

    def get_exitnum(self, lst):
        """
        Set number of exit menu choice to 9, 99, or 999 based on length of menu options.

        :param lst: List of menu options (list)
        :return exitnum: Number of exit menu choice (str)
        """
        exitnum = '9'
        if len(lst) > 8:
            exitnum = '99'
        if len(lst) > 98:
            exitnum = '999'
        return exitnum

    def get_replacements(self, prompt, replacements):
        """
        Use indirect notation to create prompt, replacing {1} with what's in dict['1'].

        :param prompt: Input string with '{}' (str)
        :param replacements: dict containing keys to replace with values (dict)
        :return prompt: Modified prompt (str)
        """
        for k in replacements:
            path = self.replace_at_symbol(replacements[k])
            replacement_text = self.d[path]
            prompt = prompt.replace('{' + k + '}', replacement_text)
        return prompt

    def replace_at_symbol(self, txt):
        """
        Use '@' notation to replace substrings with corresponding runtime variables.

        :param txt: Input text ex. - 'projects.@active_project' (str)
        :return path: Modified dot path to resource (str)
        """
        words = txt.split('.')
        altered_words = []
        for word in words:
            if '@' in word:
                altered_words.append(getattr(self, word.replace('@', '')))
            else:
                altered_words.append(word)
        path = '.'.join(altered_words)
        return path

    def roll_name(self, forbidden_list):
        """
        Get a new resource name that does not already exist in self.d['resources']

        :param forbidden_list: list of entries the random name cannot be.
        """
        x = ''.join(random.choices(string.ascii_letters + string.digits, k=6))
        while x in forbidden_list:
            x = ''.join(random.choices(string.ascii_letters + string.digits, k=6))
        return x

    # Resource Methods
    def get_project(self, text):
        """
        Find a project by search text, first containing match.

        :param text: Input text (str)
        :return: Project dict itself, not path (dict), or '' if error.
        """
        for k in self.d['project_paths']:
            if text.upper() in self.d[k]['text'].upper():
                return k
        print('Error: text ' + text + ' not found.')
        return ''

    def get_story(self, project, text):
        """
        Find a story by search text, first match.
        Function must include project, because every project contains similarly named stories  # TODO: Account for multiples.

        :param project: project path (str)
        :param text: Input text (str)
        :return: Story dict itself, not path (dict), or '' if error.
        """
        for k in self.d['storys']:
            if text.upper() in self.d[k]['text'].upper() and project in k:
                return k
        print('Error: text ' + text + ' not found.')
        return ''

    def get_task(self, text):
        """
        Find a task by search text, first match.

        :param text: Input text (str)
        :return: Task dict itself, not path (dict), or '' if error
        """
        for k in self.d['tasks']:
            if text.upper() in self.d[k]['text'].upper():
                return k
        print('Error: text ' + text + ' not found.')
        return ''

    def add_project(self, text):
        """
        Add project resource.

        :param text: Project text (str)
        """
        name = self.roll_name(self.d['resources'])
        time_created = datetime.now().strftime(self.time_format)
        self.d['projects'][name] = {
            'type': 'project',
            'text': text,
            'storys': {},
            'time_created': time_created
        }
        self.d['project_paths'].insert(0, '.'.join(['projects', name]))
        self.d['resources'].insert(0, name)
        self.add_story('.'.join(['projects', name]), 'Backlog')
        self.d['log'].append('Added Project ' + '.'.join(['projects', name]) + ' at ' + time_created)

    def add_story(self, project, text):
        """
        Add story resource.

        :param project: Parent project dot path (str)
        :param text: Story text (str)
        """
        name = self.roll_name(self.d['resources'])
        time_created = datetime.now().strftime(self.time_format)
        project = project.split('.')[-1]
        self.d['projects.' + project]['storys'][name] = {
            'type': 'story',
            'project': project,
            'text': text,
            'tasks': {},
            'time_created': time_created
        }
        self.d['resources'].insert(0, name)
        self.d['storys'].insert(0, '.'.join(['projects', project, 'storys', name]))
        self.d['log'].append(
            'Added Story ' + '.'.join(['projects', project, 'storys', name]) + ' at ' + time_created)

    def add_task(self, project, story, text, status='todo'):
        """
        Add task resource.

        :param project: Parent project dot path (str)
        :param story: Parent story dot path (str)
        :param text: Task text (str)
        :param status: Optional status (str)
        """
        story = story.split('.')[-1]
        project = project.split('.')[-1]
        name = self.roll_name(self.d['resources'])
        time_created = datetime.now().strftime(self.time_format)
        self.d['projects.' + project]['storys'][story]['tasks'][name] = {
            'type': 'task',
            'project': project,
            'story': story,
            'text': text,
            'status': status,
            'time_created': time_created,
            'notes': ''
        }
        self.d['resources'].insert(0, name)
        self.d['tasks'].insert(0, '.'.join(['projects', project, 'storys', story, 'tasks', name]))
        self.d['log'].append('Added Task ' + '.'.join(['projects', project, 'storys', story, 'tasks', name]) +
                             ' at ' + time_created)

    def del_project(self, project_id):
        """
        Remove project resource and children.

        :param project_id: Project dot path (str)
        """
        project_story_names = self.d[project_id]['storys']
        project_task_names = []
        for k in project_story_names:
            project_task_names.extend(self.d[project_id]['storys'][k]['tasks'])
        project_task_paths = []
        for k in project_task_names:
            project_task_paths.append([j for j in self.d['tasks'] if k in j][0])
        project_story_paths = []
        for k in project_story_names:
            project_story_paths.append([j for j in self.d['storys'] if k in j][0])
        for k in project_task_paths:
            self.del_task(k)
        for k in project_story_paths:
            self.del_story(k)
        del self.d[project_id]
        self.d['project_paths'].remove(project_id)
        self.d['resources'].remove(project_id.split('.')[-1])
        time_deleted = datetime.now().strftime(self.time_format)
        self.d['log'].append('Deleted Project ' + project_id + ' at ' + time_deleted)

    def del_story(self, full_story_id):
        """
        Remove story resource and children.

        :param full_story_id: story dot path (str)
        """
        story_task_names = self.d[full_story_id]['tasks']
        story_task_paths = []
        for k in story_task_names:
            story_task_paths.append([j for j in self.d['tasks'] if k in j][0])
        for k in story_task_paths:
            self.del_task(k)
        del self.d[full_story_id]
        self.d['storys'].remove(full_story_id)
        self.d['resources'].remove(full_story_id.split('.')[-1])
        time_deleted = datetime.now().strftime(self.time_format)
        self.d['log'].append('Deleted Story ' + full_story_id + ' at ' + time_deleted)

    def del_task(self, full_task_id):
        """
        Remove task resource.

        :param full_task_id: Task dot path (str)
        """
        del self.d[full_task_id]
        self.d['tasks'].remove(full_task_id)
        self.d['resources'].remove(full_task_id.split('.')[-1])
        time_deleted = datetime.now().strftime(self.time_format)
        self.d['log'].append('Deleted Task ' + full_task_id + ' at ' + time_deleted)

    # Output Methods
    def print_menu(self, option_list, exitnum):
        """
        Print menu into console.

        :param option_list: List of menu options (list)
        :param exitnum: Exit option (str)
        """
        for i, k in enumerate(option_list):
            print(str(i) + ': ' + k)
        print(exitnum + ': ' + 'Save and Exit')

    def print_scrum_board(self, target_projects):
        """
        Print scrum board into console showing all tasks sorted into status.

        :param target_projects: Project paths to be included in board (str list)
        """
        for k in target_projects:
            table = self.get_project_table(k)
            length_list = [len(element) for row in table for element in row]
            column_width = max(length_list)
            for row in table:
                row = "|".join(element.ljust(column_width + 2) for element in row)
                print(row)
            print('')

    def print_to_excel(self, target_projects, filename):
        """
        Print scrum board into Excel, one project per sheet.
        
        :param filename: .xlsx filename (str)
        :param target_projects: Project paths to be included in board (str list)
        """
        wb = Workbook()
        ws = wb.worksheets[0]
        corner = [1, 1]
        exf = self.d['excel_fmt']
        titlefont = Font(
            name=exf['HeaderFont'],
            size=int(exf['TitleFontSize']),
            bold=True,
            )
        headerfont = Font(
            name=exf['HeaderFont'],
            size=int(exf['HeaderFontSize']),
            bold=True,
            )
        textfont = Font(
            name=exf['HeaderFont'],
            size=int(exf['TextFontSize']),
            bold=True,
            )
        alignment = Alignment(
            horizontal='center',
            vertical='bottom',
            wrap_text=True)
        bd = Side(style='thick', color='000000')
        postit = NamedStyle(name='postit')
        postit.border = Border(right=bd, bottom=bd)
        for k in target_projects:
            table = self.get_project_table(k)
            projectname = table[0][0]
            wb.create_sheet(projectname[0:30])
            ws = wb[projectname[0:30]]
            for r in range(0, len(table)):
                for c in range(0, len(table[0])):
                    cell = ws.cell(row = (corner[1] + r)*2, column = (corner[0] + c)*2)
                    cell.value = table[r][c]
                    cell.alignment = alignment
                    if [r, c] == [0, 0]:
                        thisfont = titlefont
                    elif r == 1:
                        thisfont = headerfont
                    else:
                        thisfont = textfont
                    if c == 0:
                        ws.column_dimensions[get_column_letter((corner[0] + c)*2)].width = exf['StoriesWidth']
                        ws.column_dimensions[get_column_letter((corner[0] + c)*2+1)].width = exf['GapWidth']
                    else:
                        ws.column_dimensions[get_column_letter((corner[0] + c)*2)].width = exf['ColumnWidth']
                        ws.column_dimensions[get_column_letter((corner[0] + c)*2+1)].width = exf['GapWidth']
                    ws.row_dimensions[(corner[1] + r)*2].height = exf['RowHeight']
                    ws.row_dimensions[(corner[1] + r)*2+1].height = exf['GapHeight']
                    if cell.value != "":
                        cell.style = postit
                        cell.font = thisfont
                        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        if c == 0:
                            cell.fill = PatternFill('solid', fgColor=exf['StoriesColor'])
                        elif c == 1:
                            cell.fill = PatternFill('solid', fgColor=exf['ToDoColor'])
                        elif c == 2:
                            cell.fill = PatternFill('solid', fgColor=exf['InProgressColor'])
                        elif c == 3:
                            cell.fill = PatternFill('solid', fgColor=exf['ReviewColor'])
                        elif c == 4:
                            cell.fill = PatternFill('solid', fgColor=exf['BlockedColor'])
                        elif c == 5:
                            cell.fill = PatternFill('solid', fgColor=exf['CompleteColor'])
        del wb['Sheet']
        wb.save(filename)

    def get_project_table(self, target_project):
        """
        Format a project into a scrum table for use in print_scrum_board and print_to_excel

        :param target_project: Project path to make into table (str)
        :return: scrum table (list of lists)
        """
        table = []
        storys = [n for n in self.d['storys'] if target_project in n]
        table.append([self.d[target_project]['text'], '', '', '', '', ''])
        table.append(['STORY', 'TODO', 'IN PROGRESS', 'IN REVIEW', 'BLOCKED', 'COMPLETE'])
        for s in storys:
            tasks = [n for n in self.d['tasks'] if s in n]
            todo = [self.d[n]['text'] + ' ' + self.d[n]['notes'] for n in tasks if self.d[n]['status'].upper() == 'TODO']
            inprogress = [self.d[n]['text'] + ' ' + self.d[n]['notes'] for n in tasks if self.d[n]['status'].upper() == 'IN PROGRESS']
            inreview = [self.d[n]['text'] + ' ' + self.d[n]['notes'] for n in tasks if self.d[n]['status'].upper() == 'REVIEW']
            blocked = [self.d[n]['text'] + ' ' + self.d[n]['notes'] for n in tasks if self.d[n]['status'].upper() == 'BLOCKED']
            complete = [self.d[n]['text'] + ' ' + self.d[n]['notes'] for n in tasks if self.d[n]['status'].upper() == 'COMPLETE']
            story_depth = max(len(x) for x in [todo, inprogress, inreview, blocked, complete])
            story_depth = max(story_depth, 1)
            story_col = [self.d[s]['text']] + [''] * (story_depth - 1)
            for n in [todo, inprogress, inreview, blocked, complete]:
                while len(n) < story_depth:
                    n.append('')
            for n in range(0, story_depth):
                table.append([str(j[n]) for j in [story_col, todo, inprogress, inreview, blocked, complete]])
        return table      

    def write_esp_data(self, data_file):
        """
        Output modified dict data into JSON file.

        :param data_file: Target json file (str)
        """
        with open(data_file, 'w') as f:
            d = dict(self.d)  # Un-benedict the dict
            json.dump(d, f, indent=4, sort_keys=True)


A = EspApp('esp_scrum_data.json', 'esp_scrum_data.json')  # Input, output json files.
A.run()
